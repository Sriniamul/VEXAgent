"""
GitHub Alerts Excel & PDF exporter.

Exports pipeline run data (Dependabot, code-scanning, secret-scanning alerts)
from the in-memory DashboardStore — or raw GitHub API alert payloads — into:
  - Excel (.xlsx) via openpyxl
  - PDF  (.pdf)  via reportlab
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Raw GitHub alert → export dict converters
# ---------------------------------------------------------------------------

def _convert_dependabot(alert: dict[str, Any], repo: str) -> dict[str, Any]:
    """Convert a raw Dependabot API alert into the export dict format."""
    advisory = alert.get("security_advisory") or {}
    vuln = alert.get("security_vulnerability") or {}
    dep = alert.get("dependency") or {}
    pkg = dep.get("package") or {}
    identifiers = {i["type"]: i["value"] for i in (advisory.get("identifiers") or [])}
    return {
        "timestamp": alert.get("created_at") or datetime.now(timezone.utc).isoformat(),
        "repo": repo,
        "alert_id": alert.get("number"),
        "alert_type": "dependabot",
        "package_name": pkg.get("name", "unknown"),
        "cve_id": identifiers.get("CVE") or identifiers.get("GHSA") or advisory.get("ghsa_id"),
        "severity": (vuln.get("severity") or advisory.get("severity") or "medium").lower(),
        "scope": (dep.get("scope") or "").lower(),
        "html_url": alert.get("html_url", ""),
        # Secret-scanning-specific fields (blank for Dependabot)
        "secret_type": "",
        "resolution": "",
        "validity": "",
        "push_protection_bypassed": "",
        # Analysis pipeline fields
        "decision": f"state:{alert.get('state', 'open')}",
        "vex_status": "",
        "jira_key": None,
        "epss_score": None,
        "reachable": None,
        "errors": [],
        "duration_ms": None,
        "justification": "",
    }


def _convert_code_scanning(alert: dict[str, Any], repo: str) -> dict[str, Any]:
    """Convert a raw code-scanning API alert into the export dict format."""
    rule = alert.get("rule") or {}
    tool = alert.get("tool") or {}
    return {
        "timestamp": alert.get("created_at") or datetime.now(timezone.utc).isoformat(),
        "repo": repo,
        "alert_id": alert.get("number"),
        "alert_type": "code_scanning",
        "package_name": f"{tool.get('name', 'scanner')}: {rule.get('id', '?')}",
        "cve_id": rule.get("id"),
        "severity": (
            rule.get("severity")
            or rule.get("security_severity_level")
            or "medium"
        ).lower(),
        "scope": "",
        "html_url": alert.get("html_url", ""),
        # Secret-scanning-specific fields (blank for Code Scanning)
        "secret_type": "",
        "resolution": "",
        "validity": "",
        "push_protection_bypassed": "",
        # Analysis pipeline fields
        "decision": f"state:{alert.get('state', 'open')}",
        "vex_status": "",
        "jira_key": None,
        "epss_score": None,
        "reachable": None,
        "errors": [],
        "duration_ms": None,
        "justification": "",
    }


def _convert_secret_scanning(alert: dict[str, Any], repo: str) -> dict[str, Any]:
    """Convert a raw secret-scanning API alert into the export dict format."""
    bypassed = alert.get("push_protection_bypassed")
    return {
        "timestamp": alert.get("created_at") or datetime.now(timezone.utc).isoformat(),
        "repo": repo,
        "alert_id": alert.get("number"),
        "alert_type": "secret_scanning",
        "package_name": alert.get("secret_type_display_name") or alert.get("secret_type", "unknown"),
        "cve_id": f"SECRET-{alert.get('secret_type', 'unknown')}",
        "severity": "high",
        "scope": "",
        "html_url": alert.get("html_url", ""),
        # Secret-scanning-specific detail fields
        "secret_type": alert.get("secret_type", ""),
        "resolution": alert.get("resolution") or "",
        "validity": (alert.get("validity") or "unknown").replace("_", " ").title(),
        "push_protection_bypassed": "Yes" if bypassed else ("No" if bypassed is False else ""),
        # Analysis pipeline fields
        "decision": f"state:{alert.get('state', 'open')}",
        "vex_status": "",
        "jira_key": None,
        "epss_score": None,
        "reachable": None,
        "errors": [],
        "duration_ms": None,
        "justification": "",
    }


def raw_alerts_to_dicts(
    dependabot: list[dict[str, Any]],
    code_scanning: list[dict[str, Any]],
    secret_scanning: list[dict[str, Any]],
    repo: str,
) -> list[dict[str, Any]]:
    """Convert raw GitHub API alert payloads into the flat dict list used by
    :func:`export_excel` and :func:`export_pdf`."""
    rows: list[dict[str, Any]] = []
    for a in dependabot:
        try:
            rows.append(_convert_dependabot(a, repo))
        except Exception as exc:
            logger.warning("Could not convert Dependabot alert: %s", exc)
    for a in code_scanning:
        try:
            rows.append(_convert_code_scanning(a, repo))
        except Exception as exc:
            logger.warning("Could not convert code-scanning alert: %s", exc)
    for a in secret_scanning:
        try:
            rows.append(_convert_secret_scanning(a, repo))
        except Exception as exc:
            logger.warning("Could not convert secret-scanning alert: %s", exc)
    return rows

# ---------------------------------------------------------------------------
# Column definitions shared by both formats
# ---------------------------------------------------------------------------

COLUMNS = [
    ("Timestamp",    "timestamp",    18),
    ("Repository",   "repo",         30),
    ("Alert ID",     "alert_id",      9),
    ("Alert Type",   "alert_type",   14),
    ("Scope",        "scope",         12),
    ("Package",      "package_name", 28),
    ("CVE / ID",     "cve_id",       20),
    ("Severity",     "severity",     11),
    ("Alert URL",    "html_url",     42),
    # Secret-scanning-specific detail columns
    ("Secret Type",              "secret_type",               28),
    ("Resolution",               "resolution",                15),
    ("Validity",                 "validity",                  12),
    ("Push Protection Bypassed", "push_protection_bypassed",  12),
    # Analysis pipeline columns
    ("Decision",       "decision",       22),
    ("VEX Status",     "vex_status",     18),
    ("EPSS Score",     "epss_score",     10),
    ("Reachable",      "reachable",       9),
    ("Jira Ticket",    "jira_key",       13),
    ("Justification",  "justification",  55),
    ("Errors",         "errors",         30),
]

# Keys that only contain meaningful data after the VEX analysis pipeline has
# run (i.e. from --simulate or import-and-analyse).  When exporting raw
# GitHub alerts these columns are omitted automatically.
_ANALYSIS_KEYS = {"decision", "vex_status", "epss_score", "reachable", "jira_key", "justification", "errors"}

# Secret-scanning-specific columns — shown only when the export contains at
# least one secret_scanning alert.
_SECRET_KEYS = {"secret_type", "resolution", "validity", "push_protection_bypassed"}


def _has_analysis_data(runs: list[dict[str, Any]]) -> bool:
    """Return True when at least one row contains non-empty analysis data."""
    for r in runs:
        for k in _ANALYSIS_KEYS:
            v = r.get(k)
            if v is None or v == "" or v == []:
                continue
            # "state:open" is a raw-alert placeholder, not real analysis
            if k == "decision" and isinstance(v, str) and v.startswith("state:"):
                continue
            return True
    return False


def _has_secret_alerts(runs: list[dict[str, Any]]) -> bool:
    """Return True when any row is a secret_scanning alert."""
    return any(r.get("alert_type") == "secret_scanning" for r in runs)


def _select_columns(
    columns: list[tuple[str, str, int]],
    runs: list[dict[str, Any]],
) -> list[tuple[str, str, int]]:
    """Filter out analysis-only and secret-only columns when not applicable."""
    hide: set[str] = set()
    if not _has_analysis_data(runs):
        hide |= _ANALYSIS_KEYS
    if not _has_secret_alerts(runs):
        hide |= _SECRET_KEYS
    if not hide:
        return columns
    return [c for c in columns if c[1] not in hide]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt(value: Any, key: str) -> str:
    """Format a cell value for display."""
    if value is None:
        return ""
    if key == "epss_score":
        return f"{value:.4f}" if isinstance(value, (int, float)) else str(value)
    if key == "reachable":
        return "Yes" if value else ("No" if value is False else "")
    if key == "errors":
        if isinstance(value, list):
            return "; ".join(value) if value else ""
        return str(value)
    if key == "timestamp":
        # Shorten ISO timestamp to readable form
        try:
            dt = datetime.fromisoformat(str(value))
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(value)
    return str(value)


_SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}


def _sort_runs(runs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sort runs by severity (critical first), then by timestamp descending."""
    return sorted(
        runs,
        key=lambda r: (
            _SEVERITY_ORDER.get((r.get("severity") or "").lower(), 99),
            r.get("timestamp", ""),
        ),
    )


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(runs: list[dict[str, Any]]) -> bytes:
    """
    Generate an Excel workbook with all alert data.

    Returns the .xlsx file content as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Determine which columns to use (skip analysis columns for raw alerts)
    active_columns = _select_columns(COLUMNS, runs)

    # ── Summary sheet ─────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    label_font = Font(name="Calibri", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws_summary["A1"] = "VEX Agent — GitHub Security Alerts Report"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:D1")

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws_summary["A2"] = f"Generated: {now_str}"
    ws_summary["A2"].font = Font(name="Calibri", italic=True, color="666666")

    # Compute stats
    sorted_runs = _sort_runs(runs)
    total = len(sorted_runs)
    by_type: dict[str, int] = {}
    by_severity: dict[str, int] = {}
    by_decision: dict[str, int] = {}
    for r in sorted_runs:
        t = r.get("alert_type", "unknown")
        by_type[t] = by_type.get(t, 0) + 1
        s = r.get("severity", "unknown")
        by_severity[s] = by_severity.get(s, 0) + 1
        d = r.get("decision", "unknown")
        by_decision[d] = by_decision.get(d, 0) + 1

    row = 4
    ws_summary.cell(row=row, column=1, value="Total Alerts").font = label_font
    ws_summary.cell(row=row, column=2, value=total)
    row += 2

    ws_summary.cell(row=row, column=1, value="By Alert Type").font = label_font
    row += 1
    for k, v in sorted(by_type.items()):
        ws_summary.cell(row=row, column=1, value=f"  {k}")
        ws_summary.cell(row=row, column=2, value=v)
        row += 1
    row += 1

    ws_summary.cell(row=row, column=1, value="By Severity").font = label_font
    row += 1
    for k in ["critical", "high", "medium", "low", "informational"]:
        if k in by_severity:
            ws_summary.cell(row=row, column=1, value=f"  {k}")
            ws_summary.cell(row=row, column=2, value=by_severity[k])
            row += 1
    row += 1

    has_analysis = _has_analysis_data(runs)
    if has_analysis:
        ws_summary.cell(row=row, column=1, value="By Decision").font = label_font
        row += 1
        for k, v in sorted(by_decision.items()):
            ws_summary.cell(row=row, column=1, value=f"  {k}")
            ws_summary.cell(row=row, column=2, value=v)
            row += 1

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 15

    # ── Severity colour mapping ───────────────────────────────────────
    severity_fills = {
        "critical": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "high":     PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "medium":   PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "low":      PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }
    severity_fonts = {
        "critical": Font(name="Calibri", bold=True, color="FFFFFF"),
        "high":     Font(name="Calibri", bold=True, color="FFFFFF"),
        "medium":   Font(name="Calibri", bold=True, color="000000"),
        "low":      Font(name="Calibri", color="000000"),
    }

    # ── All Alerts sheet ──────────────────────────────────────────────
    ws_all = wb.create_sheet("All Alerts")

    # Header row
    for col_idx, (label, _key, width) in enumerate(active_columns, start=1):
        cell = ws_all.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws_all.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, run in enumerate(sorted_runs, start=2):
        for col_idx, (_label, key, _w) in enumerate(active_columns, start=1):
            val = run.get(key)
            cell = ws_all.cell(row=row_idx, column=col_idx, value=_fmt(val, key))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(key in ("errors", "justification")))
            # Colour severity column
            if key == "severity":
                sev = (val or "").lower()
                if sev in severity_fills:
                    cell.fill = severity_fills[sev]
                    cell.font = severity_fonts.get(sev, Font())

    # Auto-filter
    if sorted_runs:
        ws_all.auto_filter.ref = f"A1:{get_column_letter(len(active_columns))}{len(sorted_runs) + 1}"

    # ── Per-type sheets ───────────────────────────────────────────────
    type_names = {"dependabot": "Dependabot", "code_scanning": "Code Scanning", "secret_scanning": "Secret Scanning"}
    for type_key, sheet_name in type_names.items():
        type_runs = [r for r in sorted_runs if r.get("alert_type") == type_key]
        if not type_runs:
            continue
        ws = wb.create_sheet(sheet_name)
        for col_idx, (label, _key, width) in enumerate(active_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx, run in enumerate(type_runs, start=2):
            for col_idx, (_label, key, _w) in enumerate(active_columns, start=1):
                val = run.get(key)
                cell = ws.cell(row=row_idx, column=col_idx, value=_fmt(val, key))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=(key in ("errors", "justification")))
                if key == "severity":
                    sev = (val or "").lower()
                    if sev in severity_fills:
                        cell.fill = severity_fills[sev]
                        cell.font = severity_fonts.get(sev, Font())
        if type_runs:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(active_columns))}{len(type_runs) + 1}"

    # Serialise
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def export_pdf(runs: list[dict[str, Any]]) -> bytes:
    """
    Generate a PDF report with alert data.

    Returns the .pdf file content as bytes.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    )

    sorted_runs = _sort_runs(runs)
    buf = io.BytesIO()

    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor("#1F4E79"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"],
        fontSize=9, textColor=colors.grey, spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "SectionHeader", parent=styles["Heading2"],
        fontSize=13, textColor=colors.HexColor("#1F4E79"),
        spaceBefore=12, spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "CellText", parent=styles["Normal"],
        fontSize=7, leading=9,
    )

    elements: list = []

    # Title
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    elements.append(Paragraph("VEX Agent — GitHub Security Alerts Report", title_style))
    elements.append(Paragraph(f"Generated: {now_str}", subtitle_style))

    # ── Summary section ───────────────────────────────────────────────
    total = len(sorted_runs)
    by_severity: dict[str, int] = {}
    by_type: dict[str, int] = {}
    by_decision: dict[str, int] = {}
    for r in sorted_runs:
        by_severity[r.get("severity", "unknown")] = by_severity.get(r.get("severity", "unknown"), 0) + 1
        by_type[r.get("alert_type", "unknown")] = by_type.get(r.get("alert_type", "unknown"), 0) + 1
        by_decision[r.get("decision", "unknown")] = by_decision.get(r.get("decision", "unknown"), 0) + 1

    summary_data = [
        ["Metric", "Count"],
        ["Total Alerts", str(total)],
    ]
    for sev in ["critical", "high", "medium", "low", "informational"]:
        if sev in by_severity:
            summary_data.append([f"  {sev.title()}", str(by_severity[sev])])
    summary_data.append(["", ""])
    for t, c in sorted(by_type.items()):
        summary_data.append([f"  {t}", str(c)])

    elements.append(Paragraph("Summary", section_style))
    summary_table = Table(summary_data, colWidths=[150, 60])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 10 * mm))

    # ── Severity colour lookup for table cells ────────────────────────
    severity_colors = {
        "critical": colors.HexColor("#FF4444"),
        "high":     colors.HexColor("#FF8C00"),
        "medium":   colors.HexColor("#FFD700"),
        "low":      colors.HexColor("#90EE90"),
    }

    # ── Alert detail tables (per type) ────────────────────────────────
    # Use a subset of columns for PDF (limited width)
    _pdf_all_columns = [
        ("Alert ID",    "alert_id",                   30),
        ("Scope",       "scope",                       40),
        ("Package",     "package_name",                80),
        ("CVE / ID",    "cve_id",                      65),
        ("Severity",    "severity",                    45),
        # Secret-scanning detail columns
        ("Secret Type",  "secret_type",                55),
        ("Resolution",   "resolution",                 40),
        ("Validity",     "validity",                   35),
        ("Push Prot.",   "push_protection_bypassed",   30),
        # Analysis pipeline columns
        ("Decision",       "decision",                 75),
        ("VEX Status",     "vex_status",               55),
        ("EPSS",           "epss_score",               35),
        ("Reachable",      "reachable",                35),
        ("Jira",           "jira_key",                 45),
        ("Justification",  "justification",           140),
    ]
    # Dynamically filter columns using the same logic as Excel
    pdf_columns = _select_columns(_pdf_all_columns, runs)
    col_widths = [c[2] for c in pdf_columns]

    type_names = {
        "dependabot": "Dependabot Alerts",
        "code_scanning": "Code Scanning Alerts",
        "secret_scanning": "Secret Scanning Alerts",
    }

    for type_key, section_title in type_names.items():
        type_runs = [r for r in sorted_runs if r.get("alert_type") == type_key]
        if not type_runs:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(f"{section_title} ({len(type_runs)})", section_style))

        # Header row
        header = [Paragraph(f"<b>{c[0]}</b>", cell_style) for c in pdf_columns]
        table_data = [header]

        for run in type_runs:
            row_cells = []
            for _label, key, _w in pdf_columns:
                val = _fmt(run.get(key), key)
                # Convert newlines to <br/> for multi-line fields in PDF
                if key == "justification" and val:
                    # Escape XML special chars, then convert newlines to <br/>
                    val = (val
                           .replace("&", "&amp;")
                           .replace("<", "&lt;")
                           .replace(">", "&gt;")
                           .replace("\n", "<br/>"))
                row_cells.append(Paragraph(val, cell_style))
            table_data.append(row_cells)

        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Build style commands
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]

        # Colour severity cells
        sev_col_idx = next(i for i, c in enumerate(pdf_columns) if c[1] == "severity")
        for row_idx, run in enumerate(type_runs, start=1):
            sev = (run.get("severity") or "").lower()
            if sev in severity_colors:
                style_cmds.append(
                    ("BACKGROUND", (sev_col_idx, row_idx), (sev_col_idx, row_idx), severity_colors[sev])
                )

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

    # Build
    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()
