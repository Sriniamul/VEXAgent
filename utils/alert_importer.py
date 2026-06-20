"""
Excel Alert Importer.

Parses an uploaded Excel file containing GitHub security alert details,
normalises each row into a :class:`NormalisedFinding`, and runs the full
VEX analysis pipeline (EPSS, metadata, reachability) for every entry.

Expected Excel columns (case-insensitive, flexible matching):
  Alert ID | Alert Type | Repository | Package | CVE / ID | Severity
  (plus optional: EPSS Score, Reachable, Decision, VEX Status, Jira Ticket, Errors)

Any sheet can contain alert rows — the importer scans the first sheet
with a recognisable header row, or falls back to the first sheet.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any, Optional

from models.vex_models import NormalisedFinding, Severity

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column name → canonical key mapping
# ---------------------------------------------------------------------------

_COLUMN_MAP: dict[str, str] = {
    "alert id":       "alert_id",
    "alertid":        "alert_id",
    "alert_id":       "alert_id",
    "id":             "alert_id",
    "alert type":     "alert_type",
    "alerttype":      "alert_type",
    "alert_type":     "alert_type",
    "type":           "alert_type",
    "repository":     "repo",
    "repo":           "repo",
    "package":        "package_name",
    "package_name":   "package_name",
    "package name":   "package_name",
    "cve / id":       "cve_id",
    "cve/id":         "cve_id",
    "cve_id":         "cve_id",
    "cve":            "cve_id",
    "ghsa":           "cve_id",
    "severity":       "severity",
    "epss score":     "epss_score",
    "epss":           "epss_score",
    "epss_score":     "epss_score",
    "reachable":      "reachable",
    "decision":       "decision",
    "vex status":     "vex_status",
    "vex_status":     "vex_status",
    "jira ticket":    "jira_key",
    "jira":           "jira_key",
    "jira_key":       "jira_key",
    "errors":         "errors",
    "timestamp":      "timestamp",
}

_SEVERITY_MAP = {
    "critical": Severity.CRITICAL,
    "high":     Severity.HIGH,
    "medium":   Severity.MEDIUM,
    "low":      Severity.LOW,
    "info":     Severity.LOW,
    "informational": Severity.LOW,
}

_TYPE_NORMALISE = {
    "dependabot":       "dependabot",
    "sca":              "dependabot",
    "dependency":       "dependabot",
    "code_scanning":    "code_scanning",
    "code scanning":    "code_scanning",
    "codescan":         "code_scanning",
    "sast":             "code_scanning",
    "secret_scanning":  "secret_scanning",
    "secret scanning":  "secret_scanning",
    "secret":           "secret_scanning",
    "secrets":          "secret_scanning",
}


def _normalise_col(name: str) -> str | None:
    """Map a free-form column header to a canonical key."""
    clean = re.sub(r"\s+", " ", str(name).strip().lower())
    return _COLUMN_MAP.get(clean)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Parse an Excel workbook and return a list of row dicts keyed by
    canonical field names.

    Scans every sheet looking for a header row that contains at least
    ``alert_id`` and ``severity``.  Returns rows from the first matching
    sheet, or raises ``ValueError`` if none found.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    for ws in wb.worksheets:
        header_row: list[str | None] = []
        col_map: dict[int, str] = {}

        # Try the first row as header
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
            for cell in row:
                canon = _normalise_col(cell.value or "")
                header_row.append(canon)
                if canon:
                    col_map[cell.column - 1] = canon

        # Must have at least alert_id (or id) and severity
        found_keys = set(col_map.values())
        if "alert_id" not in found_keys or "severity" not in found_keys:
            continue

        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record: dict[str, Any] = {}
            for idx, val in enumerate(row):
                key = col_map.get(idx)
                if key:
                    record[key] = val
            # Skip empty rows
            if not record.get("alert_id") and not record.get("package_name"):
                continue
            rows.append(record)

        wb.close()
        logger.info("Parsed %d alert rows from sheet '%s'", len(rows), ws.title)
        return rows

    wb.close()
    raise ValueError(
        "No valid alert sheet found. The Excel file must have a header row "
        "containing at least 'Alert ID' and 'Severity' columns."
    )


def rows_to_findings(
    rows: list[dict[str, Any]],
    default_repo: str = "",
    default_branch: str = "main",
) -> list[tuple[NormalisedFinding, str]]:
    """
    Convert parsed row dicts into ``(NormalisedFinding, alert_type)`` pairs.

    Rows that cannot be converted are logged and skipped.
    """
    findings: list[tuple[NormalisedFinding, str]] = []

    for i, row in enumerate(rows, start=1):
        try:
            alert_id = int(row.get("alert_id", 0) or 0)
            if not alert_id:
                logger.warning("Row %d: missing alert_id — skipping", i)
                continue

            sev_str = str(row.get("severity", "medium")).strip().lower()
            severity = _SEVERITY_MAP.get(sev_str, Severity.MEDIUM)

            raw_type = str(row.get("alert_type", "dependabot")).strip().lower()
            alert_type = _TYPE_NORMALISE.get(raw_type, "dependabot")

            repo = str(row.get("repo", "") or default_repo or "unknown/repo")
            pkg = str(row.get("package_name", "") or "unknown")
            cve_id = str(row.get("cve_id", "") or "") or None

            clone_url = f"https://github.com/{repo}.git"

            # Determine ecosystem heuristic from alert type and package name
            ecosystem = _guess_ecosystem(alert_type, pkg)

            finding = NormalisedFinding(
                alert_id=alert_id,
                repo_full_name=repo,
                repo_clone_url=clone_url,
                repo_default_branch=default_branch,
                cve_id=cve_id if cve_id and cve_id.upper().startswith("CVE-") else None,
                ghsa_id=cve_id if cve_id and not cve_id.upper().startswith("CVE-") else None,
                package_name=pkg,
                package_version="",
                package_ecosystem=ecosystem,
                vulnerable_version_range="",
                patched_version=None,
                severity=severity,
                cvss_score=None,
                manifest_path=None,
                scope=None,
                vulnerable_functions=[],
                summary=f"Imported from Excel — {alert_type} alert #{alert_id}",
                references=[],
            )
            findings.append((finding, alert_type))

        except Exception as exc:
            logger.warning("Row %d: conversion failed — %s", i, exc)

    logger.info("Converted %d / %d rows to NormalisedFinding objects", len(findings), len(rows))
    return findings


def _guess_ecosystem(alert_type: str, package_name: str) -> str:
    """Best-effort ecosystem guess from alert type and package name."""
    if alert_type == "code_scanning":
        return "source"
    if alert_type == "secret_scanning":
        return "source"

    pkg_lower = package_name.lower()
    # Common heuristics
    if any(sep in pkg_lower for sep in (".", "com.", "org.")):
        if "json" in pkg_lower and "newtonsoft" in pkg_lower:
            return "nuget"
        return "maven"
    if pkg_lower.startswith("@") or "/" in pkg_lower:
        return "npm"
    if any(kw in pkg_lower for kw in ("rack", "rails", "nokogiri", "bundler")):
        return "rubygems"
    # Default to pip for simple names
    return "pip"
